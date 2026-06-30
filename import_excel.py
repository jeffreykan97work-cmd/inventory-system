"""
Corrected Excel import — reads the 餘額 (balance) row directly
instead of re-computing from transactions (which double-counts summary rows).
"""
import sys, re, json
sys.path.insert(0, 'D:/HERMES_WORK/inventory-system')

import openpyxl
from database import init_db, get_db

XLSX = r"D:\.hermes\cache\documents\doc_d4f2867e0350_1.____________20260604.xlsx"

STANDARD_HEADERS = {'申請編號', '日期', '分類', '事項', '申請/主管簽署日期'}

def clean_name(s):
    if not s: return ""
    s = str(s).replace('\n', ' ').replace('\r', '')
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def is_header_cell(val):
    """Check if a cell value looks like a standard header (not a product name)"""
    if not val: return False
    v = clean_name(val)
    return v in STANDARD_HEADERS

def parse_sheet(ws, sheet_name):
    """
    Parse a sheet by reading product names from row 2 and balances from the 餘額 row.
    """
    # Step 1: Find where product columns start
    # Scan row 2 from left to right — first non-header cell is the first product
    product_start_col = None
    for c in range(1, ws.max_column + 1):
        val = clean_name(ws.cell(2, c).value)
        if val and not is_header_cell(val):
            product_start_col = c
            break
    
    if product_start_col is None:
        print(f"  Skipping {sheet_name}: no product headers found in row 2")
        return []
    
    # Step 2: Find the balance (餘額) row
    balance_row = None
    for r in range(ws.max_row, max(1, ws.max_row - 20), -1):
        for c in range(1, min(6, product_start_col + 2)):
            v = str(ws.cell(r, c).value or '')
            if '餘額' in v:
                balance_row = r
                break
        if balance_row:
            break
    
    if balance_row is None:
        print(f"  WARNING: {sheet_name} — no 餘額 row found, skipping")
        return []
    
    # Step 3: Read product names from row 2 and balances from balance row
    products = []
    col = product_start_col
    
    while col <= ws.max_column:
        name = clean_name(ws.cell(2, col).value)
        if not name:
            col += 1
            continue
        
        # Balance is in the 'in' (存貨) column of each product pair
        bal = ws.cell(balance_row, col).value
        try:
            qty = int(float(str(bal).replace(',', '').replace('$', '')))
        except (ValueError, TypeError):
            # For headers like 論文集(V.1) where balance col offset differs,
            # try the next column
            bal2 = ws.cell(balance_row, col + 1).value
            try:
                qty = int(float(str(bal2).replace(',', '').replace('$', '')))
            except (ValueError, TypeError):
                qty = 0
        
        products.append({
            'name': name,
            'category': sheet_name,
            'current_qty': max(0, qty),  # clamp negatives to 0 for import
        })
        
        col += 2  # Each product has 2 columns (in/out)
    
    return products


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    
    all_products = []
    
    for sheet_name in wb.sheetnames:
        if sheet_name == '分類':
            continue
        ws = wb[sheet_name]
        print(f"\n=== {sheet_name} ({ws.max_row} rows) ===")
        products = parse_sheet(ws, sheet_name)
        all_products.extend(products)
        for p in products:
            print(f"  {p['name'][:45]:45s} → {p['current_qty']:>5d}")
    
    # Deduplicate: keep max qty when same name+category appears twice (e.g. 晴雨傘 at C19 & C81)
    deduped = {}
    for p in all_products:
        key = (p['name'], p['category'])
        if key in deduped:
            deduped[key]['current_qty'] = max(deduped[key]['current_qty'], p['current_qty'])
        else:
            deduped[key] = p
    all_products = list(deduped.values())
    
    print(f"\n{'='*60}")
    print(f"Total products from Excel: {len(all_products)} (deduplicated)")
    print(f"Total current stock: {sum(p['current_qty'] for p in all_products)}")
    
    # ─── Import into SQLite ───
    init_db()
    with get_db() as db:
        # --- Step 1: Identify sheet-based category IDs ---
        sheet_cat_ids = set()
        for p in all_products:
            cat = p['category']
            exists = db.execute("SELECT id FROM categories WHERE name=?", (cat,)).fetchone()
            if not exists:
                db.execute("INSERT INTO categories (name) VALUES (?)", (cat,))
            cat_id = db.execute("SELECT id FROM categories WHERE name=?", (cat,)).fetchone()['id']
            sheet_cat_ids.add(cat_id)
        
        # --- Step 2: Reset quantities for ALL products in sheet categories ---
        # Wipe all quantities to 0 first (they're all wrong anyway)
        if sheet_cat_ids:
            placeholders = ','.join('?' * len(sheet_cat_ids))
            db.execute(
                f"UPDATE products SET quantity=0 WHERE category_id IN ({placeholders})",
                list(sheet_cat_ids)
            )
        
        # --- Step 3: Upsert products from Excel ---
        imported = 0
        updated = 0
        
        for p in all_products:
            cat_id = db.execute("SELECT id FROM categories WHERE name=?", (p['category'],)).fetchone()['id']
            
            existing = db.execute(
                "SELECT id, quantity FROM products WHERE name=? AND category_id=?",
                (p['name'], cat_id)
            ).fetchone()
            
            if existing:
                # Update quantity
                db.execute(
                    "UPDATE products SET quantity=?, updated_at=datetime('now','localtime') WHERE id=?",
                    (p['current_qty'], existing['id'])
                )
                updated += 1
            else:
                cur = db.execute(
                    "INSERT INTO products (name, category_id, quantity, unit, location) VALUES (?,?,?,?,?)",
                    (p['name'], cat_id, p['current_qty'], '件', '')
                )
                if p['current_qty'] > 0:
                    db.execute(
                        "INSERT INTO transactions (product_id, type, quantity, note) VALUES (?,?,?,?)",
                        (cur.lastrowid, 'in', p['current_qty'], 'Excel匯入初始庫存(修正)')
                    )
                imported += 1
        
        print(f"\nImported: {imported} new, Updated: {updated}")
        
        # Show final stats
        total = db.execute("SELECT COUNT(*) FROM products").fetchone()[0]
        total_qty = db.execute("SELECT SUM(quantity) FROM products").fetchone()[0] or 0
        print(f"Database: {total} products, {total_qty} total qty")


if __name__ == '__main__':
    main()
