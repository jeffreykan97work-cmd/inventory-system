"""
庫存管理系統 — FastAPI Backend
"""
from fastapi import FastAPI, HTTPException, Query, UploadFile, File, Form
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
from typing import Optional
from database import get_db, init_db
from pathlib import Path
import uvicorn, json, re, base64, urllib.request, io, csv, os

app = FastAPI(title="庫存管理系統", version="1.0.0")

# ═══════════════════════════════════════
# MODELS
# ═══════════════════════════════════════
class ProductCreate(BaseModel):
    name: str
    category_id: int = 1
    sku: Optional[str] = ""
    quantity: int = 0
    unit: str = "件"
    unit_price: float = 0
    min_stock: int = 5
    location: str = ""
    notes: str = ""
    image_url: str = ""

class ProductUpdate(BaseModel):
    name: Optional[str] = None
    category_id: Optional[int] = None
    sku: Optional[str] = None
    quantity: Optional[int] = None
    unit: Optional[str] = None
    unit_price: Optional[float] = None
    min_stock: Optional[int] = None
    location: Optional[str] = None
    notes: Optional[str] = None
    image_url: Optional[str] = None

class StockMove(BaseModel):
    quantity: int
    note: str = ""

class CategoryCreate(BaseModel):
    name: str

# ═══════════════════════════════════════
# STARTUP
# ═══════════════════════════════════════
@app.on_event("startup")
def startup():
    init_db()

# ═══════════════════════════════════════
# CATEGORIES
# ═══════════════════════════════════════
@app.get("/api/categories")
def list_categories():
    with get_db() as db:
        rows = db.execute("SELECT * FROM categories ORDER BY id").fetchall()
        return [dict(r) for r in rows]

@app.post("/api/categories")
def create_category(data: CategoryCreate):
    with get_db() as db:
        try:
            cur = db.execute("INSERT INTO categories (name) VALUES (?)", (data.name,))
            return {"id": cur.lastrowid, "name": data.name}
        except:
            raise HTTPException(400, "分類已存在")

# ═══════════════════════════════════════
# PRODUCTS
# ═══════════════════════════════════════
@app.get("/api/products")
def list_products(
    search: str = Query(""),
    category_id: Optional[int] = None,
    low_stock: bool = False,
    page: int = 1,
    per_page: int = 50
):
    with get_db() as db:
        sql = """SELECT p.*, c.name as category_name 
                 FROM products p LEFT JOIN categories c ON p.category_id=c.id WHERE 1=1"""
        params = []
        if search:
            sql += " AND (p.name LIKE ? OR p.sku LIKE ? OR p.location LIKE ?)"
            q = f"%{search}%"
            params += [q, q, q]
        if category_id:
            sql += " AND p.category_id = ?"
            params.append(category_id)
        if low_stock:
            sql += " AND p.quantity <= p.min_stock"
        
        # Count
        count = db.execute(f"SELECT COUNT(*) FROM ({sql})", params).fetchone()[0]
        # Page
        sql += " ORDER BY p.updated_at DESC LIMIT ? OFFSET ?"
        params += [per_page, (page-1)*per_page]
        rows = db.execute(sql, params).fetchall()
        return {"items": [dict(r) for r in rows], "total": count, "page": page, "per_page": per_page}

@app.get("/api/products/{product_id}")
def get_product(product_id: int):
    with get_db() as db:
        row = db.execute(
            "SELECT p.*, c.name as category_name FROM products p LEFT JOIN categories c ON p.category_id=c.id WHERE p.id=?",
            (product_id,)
        ).fetchone()
        if not row:
            raise HTTPException(404, "產品不存在")
        return dict(row)

@app.post("/api/products")
def create_product(data: ProductCreate):
    with get_db() as db:
        cur = db.execute("""
            INSERT INTO products (name,category_id,sku,quantity,unit,unit_price,min_stock,location,notes)
            VALUES (?,?,?,?,?,?,?,?,?)
        """, (data.name, data.category_id, data.sku or "", data.quantity, data.unit,
              data.unit_price, data.min_stock, data.location, data.notes))
        pid = cur.lastrowid
        if data.quantity != 0:
            db.execute("INSERT INTO transactions (product_id,type,quantity,note) VALUES (?,?,?,?)",
                       (pid, 'in' if data.quantity > 0 else 'out', abs(data.quantity), '初始庫存'))
        return {"id": pid, **data.dict()}

@app.put("/api/products/{product_id}")
def update_product(product_id: int, data: ProductUpdate):
    with get_db() as db:
        fields = {k: v for k, v in data.dict().items() if v is not None}
        if not fields:
            raise HTTPException(400, "沒有要更新的欄位")
        sets = ", ".join(f"{k}=?" for k in fields)
        sql = f"UPDATE products SET {sets}, updated_at=datetime('now','localtime') WHERE id=?"
        db.execute(sql, list(fields.values()) + [product_id])
        return {"ok": True}

@app.delete("/api/products/{product_id}")
def delete_product(product_id: int):
    with get_db() as db:
        db.execute("DELETE FROM transactions WHERE product_id=?", (product_id,))
        db.execute("DELETE FROM products WHERE id=?", (product_id,))
        return {"ok": True}

# ═══════════════════════════════════════
# STOCK MOVEMENTS
# ═══════════════════════════════════════
@app.post("/api/products/{product_id}/stock-in")
def stock_in(product_id: int, data: StockMove):
    with get_db() as db:
        p = db.execute("SELECT quantity FROM products WHERE id=?", (product_id,)).fetchone()
        if not p: raise HTTPException(404, "產品不存在")
        db.execute("UPDATE products SET quantity=quantity+?, updated_at=datetime('now','localtime') WHERE id=?",
                   (data.quantity, product_id))
        db.execute("INSERT INTO transactions (product_id,type,quantity,note) VALUES (?,?,?,?)",
                   (product_id, 'in', data.quantity, data.note))
        return {"ok": True, "new_qty": p["quantity"] + data.quantity}

@app.post("/api/products/{product_id}/stock-out")
def stock_out(product_id: int, data: StockMove):
    with get_db() as db:
        p = db.execute("SELECT quantity FROM products WHERE id=?", (product_id,)).fetchone()
        if not p: raise HTTPException(404, "產品不存在")
        if p["quantity"] < data.quantity:
            raise HTTPException(400, "庫存不足")
        db.execute("UPDATE products SET quantity=quantity-?, updated_at=datetime('now','localtime') WHERE id=?",
                   (data.quantity, product_id))
        db.execute("INSERT INTO transactions (product_id,type,quantity,note) VALUES (?,?,?,?)",
                   (product_id, 'out', data.quantity, data.note))
        return {"ok": True, "new_qty": p["quantity"] - data.quantity}

# ═══════════════════════════════════════
# TRANSACTIONS
# ═══════════════════════════════════════
@app.get("/api/products/{product_id}/transactions")
def list_transactions(product_id: int, limit: int = 50):
    with get_db() as db:
        rows = db.execute(
            "SELECT * FROM transactions WHERE product_id=? ORDER BY created_at DESC LIMIT ?",
            (product_id, limit)
        ).fetchall()
        return [dict(r) for r in rows]

# ═══════════════════════════════════════
# APPLICATION FORM SUBMISSION (AI-powered)
# ═══════════════════════════════════════
class ApplicationItem(BaseModel):
    name: str
    quantity: int = 0

class ApplicationSubmit(BaseModel):
    app_type: str  # 'withdraw' or 'deposit'
    department: str = ""
    applicant_name: str = ""
    reason: str = ""
    items: list[ApplicationItem] = []

@app.post("/api/applications/submit")
def submit_application(data: ApplicationSubmit):
    """Submit application form → AI auto-identifies items → updates stock"""
    if data.app_type not in ('withdraw', 'deposit'):
        raise HTTPException(400, "類型必須是 withdraw 或 deposit")
    if not data.items or all(it.quantity <= 0 for it in data.items):
        raise HTTPException(400, "請至少填寫一項物品及數量")
    
    import re
    results = []
    
    with get_db() as db:
        for item in data.items:
            if item.quantity <= 0:
                continue
            
            norm_name = re.sub(r'[\s\-_/\(\)（）]', '', item.name).lower()
            matched_product = None
            match_type = ""
            
            # 1) Exact match: normalize both sides in Python, compare
            all_products = db.execute("SELECT * FROM products").fetchall()
            for row in all_products:
                r = dict(row)
                r_norm = re.sub(r'[\s\-_/\(\)（）]', '', r['name']).lower()
                if r_norm == norm_name:
                    matched_product = r
                    match_type = "完全匹配"
                    break
            
            # 2) Partial match via LIKE
            if not matched_product:
                q = f"%{item.name}%"
                row = db.execute(
                    "SELECT * FROM products WHERE name LIKE ? LIMIT 1", (q,)
                ).fetchone()
                if row:
                    matched_product = dict(row)
                    match_type = "近似匹配"
            
            # 3) Create new product
            if not matched_product:
                cur = db.execute(
                    "INSERT INTO products (name, category_id, quantity, unit) VALUES (?, 1, 0, '件')",
                    (item.name,)
                )
                pid = cur.lastrowid
                matched_product = {"id": pid, "name": item.name, "quantity": 0}
                match_type = "新產品"
            
            pid = matched_product["id"]
            old_qty = matched_product["quantity"]
            
            if data.app_type == 'deposit':
                db.execute("UPDATE products SET quantity=quantity+?, updated_at=datetime('now','localtime') WHERE id=?", 
                          (item.quantity, pid))
                new_qty = old_qty + item.quantity
                txn_type = 'in'
            else:
                if old_qty < item.quantity:
                    raise HTTPException(400, f"「{item.name}」庫存不足 (現有: {old_qty}, 需要: {item.quantity})")
                db.execute("UPDATE products SET quantity=quantity-?, updated_at=datetime('now','localtime') WHERE id=?", 
                          (item.quantity, pid))
                new_qty = old_qty - item.quantity
                txn_type = 'out'
            
            results.append({
                "item": item.name,
                "qty": item.quantity,
                "match": match_type,
                "product_id": pid,
                "old_qty": old_qty,
                "new_qty": new_qty,
                "action": "存入" if data.app_type == 'deposit' else "申領"
            })
        
        # Record application
        items_json = json.dumps([it.model_dump() for it in data.items])
        cur = db.execute(
            "INSERT INTO applications (app_type, department, applicant_name, reason, items_json, status) VALUES (?,?,?,?,?,'approved')",
            (data.app_type, data.department, data.applicant_name, data.reason, items_json)
        )
        app_id = cur.lastrowid
        
        # Record transactions linked to application
        for r in results:
            db.execute(
                "INSERT INTO transactions (product_id, type, quantity, note, application_id) VALUES (?,?,?,?,?)",
                (r["product_id"], txn_type, r["qty"], f"申請#{app_id}: {data.applicant_name} - {data.reason}", app_id)
            )
        
        return {
            "application_id": app_id,
            "status": "approved",
            "results": results,
            "summary": f"AI 自動處理 {len(results)} 項物品 ({'存入' if data.app_type == 'deposit' else '申領'})"
        }

@app.get("/api/applications")
def list_applications(limit: int = 20):
    with get_db() as db:
        rows = db.execute("SELECT * FROM applications ORDER BY created_at DESC LIMIT ?", (limit,)).fetchall()
        apps = []
        for r in rows:
            d = dict(r)
            d["items"] = json.loads(d["items_json"])
            del d["items_json"]
            apps.append(d)
        return apps

# ═══════════════════════════════════════
# 📸 OCR — 拍照識別紙本申請表
# ═══════════════════════════════════════
UPLOAD_DIR = Path(__file__).parent / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

OLLAMA_URL = os.environ.get("OLLAMA_URL", "http://localhost:11434/api/chat")
OLLAMA_MODEL = os.environ.get("OLLAMA_VISION_MODEL", "minicpm-v:latest")

OCR_PROMPT = """你係一個 OCR 系統。請仔細閱讀呢張「紀念品申請表」嘅照片，提取以下資訊，用 JSON 格式回覆：

{
  "app_type": "deposit 或 withdraw (表格右上角會寫「申領」或「存入」，申領=withdraw, 存入=deposit)",
  "department": "申請部門",
  "applicant_name": "申請人姓名",
  "reason": "申請事由",
  "items": [
    {"name": "物品名稱", "quantity": 數量}
  ]
}

規則：
- 只提取有填寫數量嘅項目（數量 > 0）
- 如果認唔到某個欄位，留空 string ""
- 必須係 valid JSON，唔好加解釋
- 物品名稱盡量完整提取"""

@app.post("/api/ocr/upload")
async def ocr_upload(file: UploadFile = File(...)):
    """Upload photo of paper form → AI OCR → return parsed data"""
    # Save uploaded file
    ext = Path(file.filename).suffix or ".jpg"
    save_path = UPLOAD_DIR / f"ocr_{int(__import__('time').time())}{ext}"
    content = await file.read()
    save_path.write_bytes(content)
    
    # Encode as base64
    img_b64 = base64.b64encode(content).decode()
    
    # Call Ollama vision
    payload = {
        "model": OLLAMA_MODEL,
        "messages": [{"role": "user", "content": OCR_PROMPT, "images": [img_b64]}],
        "stream": False,
        "options": {"temperature": 0.1}
    }
    
    try:
        req = urllib.request.Request(OLLAMA_URL, 
            data=json.dumps(payload).encode(), 
            headers={"Content-Type": "application/json"})
        resp = json.loads(urllib.request.urlopen(req, timeout=120).read())
        raw_text = resp["message"]["content"]
    except Exception as e:
        raise HTTPException(500, f"Ollama OCR 失敗: {str(e)}")
    
    # Parse JSON from response
    # Try to extract JSON block
    json_match = re.search(r'\{[\s\S]*\}', raw_text)
    if json_match:
        try:
            parsed = json.loads(json_match.group())
        except:
            raise HTTPException(500, f"AI 回覆格式錯誤，無法解析 JSON: {raw_text[:300]}")
    else:
        raise HTTPException(500, f"AI 回覆無 JSON: {raw_text[:300]}")
    
    # Validate and suggest matches
    suggestions = []
    with get_db() as db:
        for item in parsed.get("items", []):
            name = item.get("name", "").strip()
            qty = item.get("quantity", 0)
            if not name or qty <= 0:
                continue
            
            # Try to match against existing products
            norm = re.sub(r'[\s\-_/\(\)（）]', '', name).lower()
            all_prods = db.execute("SELECT id, name, quantity FROM products").fetchall()
            best_match = None
            for row in all_prods:
                r_norm = re.sub(r'[\s\-_/\(\)（）]', '', row['name']).lower()
                if r_norm == norm:
                    best_match = {"id": row['id'], "name": row['name'], "current_qty": row['quantity'], "confidence": "高"}
                    break
                elif norm in r_norm or r_norm in norm:
                    if not best_match:
                        best_match = {"id": row['id'], "name": row['name'], "current_qty": row['quantity'], "confidence": "中"}
            
            suggestions.append({
                "ocr_name": name,
                "ocr_qty": qty,
                "match": best_match
            })
    
    return {
        "raw_text": raw_text[:500],
        "parsed": parsed,
        "suggestions": suggestions,
        "image_path": str(save_path)
    }

@app.post("/api/ocr/confirm")
def ocr_confirm(data: ApplicationSubmit):
    """Confirm OCR results → process application"""
    return submit_application(data)

# ═══════════════════════════════════════
# ⚙️ ADMIN — 後台管理
# ═══════════════════════════════════════
@app.delete("/api/categories/{category_id}")
def delete_category(category_id: int):
    with get_db() as db:
        # Move products to uncategorized
        db.execute("UPDATE products SET category_id=1 WHERE category_id=?", (category_id,))
        db.execute("DELETE FROM categories WHERE id=? AND id > 6", (category_id,))
        return {"ok": True}

@app.post("/api/products/batch-update")
def batch_update(updates: list[dict]):
    """Batch update products: [{id, name?, category_id?, quantity?, ...}]"""
    with get_db() as db:
        count = 0
        for u in updates:
            pid = u.pop("id", None)
            if not pid or not u:
                continue
            sets = ", ".join(f"{k}=?" for k in u)
            sql = f"UPDATE products SET {sets}, updated_at=datetime('now','localtime') WHERE id=?"
            db.execute(sql, list(u.values()) + [pid])
            count += 1
        return {"ok": True, "updated": count}

@app.get("/api/export/csv")
def export_csv():
    """Export all products as CSV"""
    with get_db() as db:
        rows = db.execute("""
            SELECT p.name, c.name as category, p.sku, p.quantity, p.unit, 
                   p.unit_price, p.min_stock, p.location, p.notes
            FROM products p LEFT JOIN categories c ON p.category_id=c.id
            ORDER BY c.name, p.name
        """).fetchall()
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["名稱", "分類", "SKU", "庫存", "單位", "單價", "安全存量", "位置", "備註"])
    for r in rows:
        writer.writerow([r[i] for i in range(len(r))])
    
    output.seek(0)
    from urllib.parse import quote
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote('inventory_export.csv')}"}
    )

@app.get("/api/transactions/recent")
def recent_transactions(limit: int = 100):
    with get_db() as db:
        rows = db.execute("""
            SELECT t.*, p.name as product_name 
            FROM transactions t LEFT JOIN products p ON t.product_id=p.id
            ORDER BY t.created_at DESC LIMIT ?
        """, (limit,)).fetchall()
        return [dict(r) for r in rows]

# ═══════════════════════════════════════
# 📸 IMAGE UPLOAD
# ═══════════════════════════════════════
IMAGES_DIR = Path(__file__).parent / "static" / "images"
IMAGES_DIR.mkdir(parents=True, exist_ok=True)

@app.post("/api/upload-image")
async def upload_image(product_id: int = Form(...), file: UploadFile = File(...)):
    """Upload product image"""
    ext = Path(file.filename).suffix or ".jpg"
    fname = f"prod_{product_id}_{int(__import__('time').time())}{ext}"
    save_path = IMAGES_DIR / fname
    content = await file.read()
    save_path.write_bytes(content)
    url = f"/static/images/{fname}"
    
    with get_db() as db:
        db.execute("UPDATE products SET image_url=?, updated_at=datetime('now','localtime') WHERE id=?", (url, product_id))
    return {"ok": True, "image_url": url}

# ═══════════════════════════════════════
# 📥 EXCEL EXPORT (original format)
# ═══════════════════════════════════════
@app.get("/api/export/excel")
def export_excel_original():
    """Export in original multi-sheet Excel format matching the source file"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    with get_db() as db:
        categories = db.execute("SELECT * FROM categories ORDER BY id").fetchall()
        
        for cat in categories:
            cat_name = cat['name']
            products = db.execute(
                "SELECT * FROM products WHERE category_id=? ORDER BY name", (cat['id'],)
            ).fetchall()
            if not products:
                continue
            
            # Sheet name max 31 chars
            sheet_name = cat_name[:31]
            ws = wb.create_sheet(title=sheet_name)
            
            # Style definitions
            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Row 1: Title
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(products)*2)
            title_cell = ws.cell(1, 1, f"{cat_name}_倉庫數量表")
            title_cell.font = Font(bold=True, size=14)
            
            # Row 2: Headers — 申請編號, 日期, 分類, 事項, then products
            ws.cell(2, 1, "申請編號").font = header_font
            ws.cell(2, 2, "日期").font = header_font
            ws.cell(2, 3, "分類").font = header_font
            ws.cell(2, 4, "事項").font = header_font
            
            for i, p in enumerate(products):
                col = 5 + i * 2
                ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+1)
                cell = ws.cell(2, col, p['name'])
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
                ws.cell(2, col+1).border = thin_border
            
            # Row 3: Current stock (存貨) and blank (提取)
            ws.cell(3, 1, "現有庫存")
            for i, p in enumerate(products):
                col = 5 + i * 2
                ws.cell(3, col, p['quantity'])
                ws.cell(3, col).font = Font(bold=True, color="006100")
                ws.cell(3, col).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                ws.cell(3, col+1, "")  # 提取 column starts at 0
            
            # Row 4: Sub-headers (存貨 / 提取)
            ws.cell(4, 1, ""); ws.cell(4, 2, ""); ws.cell(4, 3, ""); ws.cell(4, 4, "")
            for i, p in enumerate(products):
                col = 5 + i * 2
                c1 = ws.cell(4, col, "存貨")
                c1.font = Font(bold=True, size=10, color="006100")
                c1.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                c2 = ws.cell(4, col+1, "提取")
                c2.font = Font(bold=True, size=10, color="9C0006") 
                c2.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # Row 5+: Transaction records
            # Get all transactions for products in this category
            prod_ids = [p['id'] for p in products]
            if prod_ids:
                placeholders = ','.join('?' * len(prod_ids))
                txns = db.execute(f"""
                    SELECT t.*, p.name as product_name
                    FROM transactions t JOIN products p ON t.product_id=p.id
                    WHERE t.product_id IN ({placeholders})
                    ORDER BY t.created_at ASC
                """, prod_ids).fetchall()
                
                row = 5
                for txn in txns:
                    ws.cell(row, 1, txn['application_id'] or '')
                    ws.cell(row, 2, txn['created_at'][:10] if txn['created_at'] else '')
                    ws.cell(row, 3, cat_name)
                    ws.cell(row, 4, txn['note'] or '')
                    
                    for i, p in enumerate(products):
                        col = 5 + i * 2
                        if txn['product_id'] == p['id']:
                            if txn['type'] == 'in':
                                ws.cell(row, col, txn['quantity'])
                            else:
                                ws.cell(row, col+1, txn['quantity'])
                    
                    row += 1
            
            # Set column widths
            ws.column_dimensions['A'].width = 14
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 14
            ws.column_dimensions['D'].width = 20
            for i in range(len(products)):
                col_letter = get_column_letter(5 + i * 2)
                ws.column_dimensions[col_letter].width = 10
                ws.column_dimensions[get_column_letter(5 + i * 2 + 1)].width = 10
    
    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    from urllib.parse import quote
    safe_name = quote("DMA_庫存表.xlsx")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{safe_name}"}
    )

# ═══════════════════════════════════════
# 🖼 CATALOG — 同事目錄 (with images)
# ═══════════════════════════════════════
@app.get("/catalog")
def catalog_page():
    return FileResponse(str(STATIC_DIR / "catalog.html"))

@app.get("/api/catalog")
def catalog_data(category_id: Optional[int] = None):
    """Return products with images for the catalog view"""
    with get_db() as db:
        sql = """SELECT p.*, c.name as category_name 
                 FROM products p LEFT JOIN categories c ON p.category_id=c.id 
                 WHERE p.quantity > 0"""
        params = []
        if category_id:
            sql += " AND p.category_id = ?"
            params.append(category_id)
        sql += " ORDER BY c.name, p.name"
        rows = db.execute(sql, params).fetchall()
        return [dict(r) for r in rows]

# ═══════════════════════════════════════
# STATS
# ═══════════════════════════════════════
@app.get("/api/stats")
def get_stats():
    with get_db() as db:
        total = db.execute("SELECT COUNT(*) FROM products").fetchone()[0]
        low = db.execute("SELECT COUNT(*) FROM products WHERE quantity <= min_stock").fetchone()[0]
        total_qty = db.execute("SELECT SUM(quantity) FROM products").fetchone()[0] or 0
        total_value = db.execute("SELECT SUM(quantity * unit_price) FROM products").fetchone()[0] or 0
        return {"total_products": total, "low_stock": low, "total_quantity": total_qty, "total_value": round(total_value, 2)}

# ═══════════════════════════════════════
# 📄 PDF 申請表生成
# ═══════════════════════════════════════
from fpdf import FPDF

# Font paths: try local bundle first, then Windows, then Render Linux fallback
_BASE = Path(__file__).parent
FONT_PATH = _BASE / "fonts" / "msjh.ttc"
FONT_BOLD_PATH = _BASE / "fonts" / "msjhbd.ttc"
_WIN_FONT_PATH = Path("C:/Windows/Fonts/msjh.ttc")
_WIN_FONT_BOLD_PATH = Path("C:/Windows/Fonts/msjhbd.ttc")
# Free CJK fallback for Render (Linux)
_NOTO_PATH = _BASE / "fonts" / "NotoSansCJKtc-Regular.otf"
_NOTO_BOLD_PATH = _BASE / "fonts" / "NotoSansCJKtc-Bold.otf"

if not FONT_PATH.exists() and _WIN_FONT_PATH.exists():
    FONT_PATH = _WIN_FONT_PATH
if not FONT_BOLD_PATH.exists() and _WIN_FONT_BOLD_PATH.exists():
    FONT_BOLD_PATH = _WIN_FONT_BOLD_PATH
# Linux fallback: use Noto Sans TC if msjh unavailable
if not FONT_PATH.exists() and _NOTO_PATH.exists():
    FONT_PATH = _NOTO_PATH
if not FONT_BOLD_PATH.exists() and _NOTO_BOLD_PATH.exists():
    FONT_BOLD_PATH = _NOTO_BOLD_PATH

class ApplicationFormPDF(FPDF):
    _has_cjk = False

    def __init__(self):
        super().__init__(orientation='P', unit='mm', format='A4')
        # Try to load CJK font
        regular_font = None
        bold_font = None
        if FONT_PATH.exists():
            regular_font = str(FONT_PATH)
        elif FONT_BOLD_PATH.exists():
            regular_font = str(FONT_BOLD_PATH)  # fallback: use bold as regular
        if FONT_BOLD_PATH.exists():
            bold_font = str(FONT_BOLD_PATH)
        if regular_font:
            self.add_font("msjh", "", regular_font)
            self._has_cjk = True
        if bold_font:
            self.add_font("msjh", "B", bold_font)
        self.set_auto_page_break(auto=True, margin=15)

    def _safe_font(self, family=None, style="", size=12):
        """Set font, falling back to built-in if CJK not available"""
        if self._has_cjk:
            self.set_font(family or "msjh", style, size)
        else:
            self.set_font("Helvetica", style, size)

    def footer(self):
        self.set_y(-15)
        self._safe_font(size=8)
        self.set_text_color(150, 150, 150)
        self.cell(0, 5, "AI 自動生成 · DMA 庫存管理系統", align="C")

    def draw_form(self, app_type: str, products: list, department="", applicant="", reason=""):
        self.add_page()
        w = 190  # usable width

        # ── Title ──
        self.set_font("msjh", "B", 20)
        self.set_text_color(30, 30, 30)
        self.cell(0, 10, "紀 念 品 申 請 表", align="C", new_x="LMARGIN", new_y="NEXT")
        self.ln(2)
        # underline
        self.set_draw_color(80, 80, 80)
        self.set_line_width(0.5)
        self.line(30, self.get_y(), 180, self.get_y())
        self.ln(6)

        # ── App type ──
        self.set_font("msjh", "", 12)
        self.set_text_color(50, 50, 50)
        self.cell(25, 8, "申請類型：")
        checked = "■" if app_type == "withdraw" else "□"
        self.cell(30, 8, f"{checked} 申領（提取）")
        checked2 = "■" if app_type == "deposit" else "□"
        self.cell(45, 8, f"{checked2} 存入（歸還）")
        self.ln(12)

        # ── Info row ──
        self.set_font("msjh", "", 12)
        self.cell(20, 8, "申請部門：")
        self.set_font("msjh", "", 12)
        x_dep = self.get_x()
        self.cell(60, 8, department if department else "_______________")
        self.cell(22, 8, "申請人：")
        self.cell(50, 8, applicant if applicant else "_______________")
        self.ln(12)

        self.cell(22, 8, "申請事由：")
        x_reason = self.get_x()
        self.cell(w - 22, 8, reason if reason else "_______________________________________________")
        self.ln(14)

        # ── Items table ──
        col_name = 110
        col_qty = 35
        col_note = 40
        row_h = 10

        # Table header
        self.set_fill_color(220, 225, 240)
        self.set_font("msjh", "B", 11)
        self.set_text_color(30, 30, 30)
        self.set_draw_color(100, 100, 100)
        self.set_line_width(0.3)

        self.cell(col_name, row_h, "  物品名稱", border=1, fill=True)
        self.cell(col_qty, row_h, "數量", border=1, fill=True, align="C")
        self.cell(col_note, row_h, "備註", border=1, fill=True, align="C")
        self.ln()

        # Table rows
        self.set_font("msjh", "", 11)
        for p in products:
            # Check if we need a page break
            if self.get_y() > 240:
                self.add_page()
                # Repeat header on new page
                self.set_fill_color(220, 225, 240)
                self.set_font("msjh", "B", 11)
                self.cell(col_name, row_h, "  物品名稱", border=1, fill=True)
                self.cell(col_qty, row_h, "數量", border=1, fill=True, align="C")
                self.cell(col_note, row_h, "備註", border=1, fill=True, align="C")
                self.ln()
                self.set_font("msjh", "", 11)

            name = p["name"][:40]  # Truncate long names
            current_qty = f"  (現存: {p['quantity']})"
            self.cell(col_name, row_h, f"  {name}", border=1)
            self.cell(col_qty, row_h, "______", border=1, align="C")
            self.cell(col_note, row_h, p.get("unit", ""), border=1, align="C")
            self.ln()

        # Extra blank rows (at least 3)
        blanks = max(3, 8 - len(products))
        for _ in range(blanks):
            if self.get_y() > 240:
                self.add_page()
            self.cell(col_name, row_h, "  ________________________", border=1)
            self.cell(col_qty, row_h, "", border=1, align="C")
            self.cell(col_note, row_h, "", border=1, align="C")
            self.ln()

        self.ln(8)

        # ── Signatures ──
        self.set_font("msjh", "", 12)
        self.set_text_color(50, 50, 50)
        self.cell(30, 8, "申請人簽名：")
        self.cell(55, 8, "______________")
        self.cell(15, 8, "日期：")
        self.cell(40, 8, "______________")
        self.ln(12)

        self.cell(30, 8, "主  管  簽名：")
        self.cell(55, 8, "______________")
        self.cell(15, 8, "日期：")
        self.cell(40, 8, "______________")
        self.ln(8)

        return self


@app.get("/api/applications/generate-pdf")
def generate_pdf(
    product_ids: str = Query("", description="Comma-separated product IDs"),
    app_type: str = Query("withdraw", description="'withdraw' or 'deposit'"),
    department: str = Query(""),
    applicant: str = Query(""),
    reason: str = Query(""),
):
    """Generate a printable PDF application form with pre-filled product names"""
    if not product_ids:
        raise HTTPException(400, "請選擇至少一件物品")

    ids = [int(x.strip()) for x in product_ids.split(",") if x.strip().isdigit()]
    if not ids:
        raise HTTPException(400, "無效的物品 ID")

    with get_db() as db:
        placeholders = ",".join("?" * len(ids))
        rows = db.execute(
            f"SELECT p.*, c.name as category_name FROM products p LEFT JOIN categories c ON p.category_id=c.id WHERE p.id IN ({placeholders}) ORDER BY c.name, p.name",
            ids,
        ).fetchall()
        products = [dict(r) for r in rows]

    if not products:
        raise HTTPException(404, "找不到指定的物品")

    pdf = ApplicationFormPDF()
    pdf.draw_form(app_type=app_type, products=products, department=department, applicant=applicant, reason=reason)

    output = io.BytesIO()
    pdf.output(output)
    output.seek(0)

    from urllib.parse import quote

    safe_name = quote("DMA_紀念品申請表.pdf")
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename*=UTF-8\'\'{safe_name}'},
    )


@app.get("/form")
def form_page():
    return FileResponse(str(STATIC_DIR / "form.html"))


# ═══════════════════════════════════════
# STATIC
# ═══════════════════════════════════════
STATIC_DIR = Path(__file__).parent / "static"
STATIC_DIR.mkdir(exist_ok=True)

app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

@app.get("/")
def index():
    return FileResponse(str(STATIC_DIR / "index.html"))

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8708)
